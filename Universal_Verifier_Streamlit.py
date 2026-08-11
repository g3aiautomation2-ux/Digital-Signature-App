import streamlit as st

import hashlib

import base64

import os

import re

import tempfile

import shutil



# PDF Imports

import PyPDF2



# Cryptography Imports

from cryptography.hazmat.primitives import serialization

from cryptography.hazmat.primitives.asymmetric import padding

from cryptography.hazmat.primitives import hashes

from cryptography.exceptions import InvalidSignature



# Excel/ODS Imports

from openpyxl import load_workbook

from odf.opendocument import load as load_ods

from odf.table import Table, TableRow, TableCell

from odf.text import P

from odf import teletype





# =========================================================

# ODS SUPPORT HELPERS

# =========================================================



def is_ods_file(file_path):

    return os.path.splitext(str(file_path))[1].lower() == ".ods"



def get_ods_cell_text(cell):

    text = teletype.extractText(cell)

    if text is not None:

        text = text.strip()

    if text:

        return text

    value = cell.getAttribute("value")

    if value is not None:

        return str(value)

    return ""



def load_ods_data(file_path):

    doc = load_ods(file_path)

    data = {}

    for table in doc.spreadsheet.getElementsByType(Table):

        sheet_name = table.getAttribute("name")

        rows = []

        for row in table.getElementsByType(TableRow):

            values = []

            for cell in row.getElementsByType(TableCell):

                repeat = cell.getAttribute("numbercolumnsrepeated")

                repeat = int(repeat) if repeat else 1

                cell_value = get_ods_cell_text(cell)

                for _ in range(repeat):

                    values.append(cell_value)

            while values and values[-1] == "":

                values.pop()

            row_repeat = row.getAttribute("numberrowsrepeated")

            row_repeat = int(row_repeat) if row_repeat else 1

            for _ in range(row_repeat):

                rows.append(values.copy())

        data[sheet_name] = rows

    return data



def load_spreadsheet_data(file_path):

    if is_ods_file(file_path):

        return load_ods_data(file_path)

    wb = load_workbook(file_path, data_only=False)

    data = {}

    for sheet_name in wb.sheetnames:

        ws = wb[sheet_name]

        rows = []

        for row in ws.iter_rows(min_row=1, min_col=1):

            row_vals = [cell.value for cell in row]

            while row_vals and (row_vals[-1] is None or str(row_vals[-1]).strip() == ""):

                row_vals.pop()

            rows.append(row_vals)

        while rows and len(rows[-1]) == 0:

            rows.pop()

        data[sheet_name] = rows

    return data



def sanitize_xlsx(input_path):

    import subprocess

    import shutil

    def _find_libreoffice_command():

        candidates = [

            shutil.which("soffice"), shutil.which("libreoffice"),

            r"C:\Program Files\LibreOffice\program\soffice.exe",

            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",

            r"C:\Program Files\OpenOffice\program\soffice.exe",

            r"C:\Program Files (x86)\OpenOffice\program\soffice.exe"

        ]

        for candidate in candidates:

            if candidate and os.path.exists(candidate): return candidate

        return None



    libreoffice_cmd = _find_libreoffice_command()

    if not libreoffice_cmd: return False

        

    temp_dir = os.path.dirname(input_path)

    temp_name = "temp_sanitize_input.xlsx"

    temp_input = os.path.join(temp_dir, temp_name)

    shutil.copy(input_path, temp_input)

    

    command1 = [libreoffice_cmd, "--headless", "--convert-to", "ods", "--outdir", temp_dir, temp_input]

    subprocess.run(command1, capture_output=True, text=True, check=False)

    

    temp_ods = os.path.join(temp_dir, "temp_sanitize_input.ods")

    if not os.path.exists(temp_ods): return False

        

    command2 = [libreoffice_cmd, "--headless", "--convert-to", "xlsx", "--outdir", temp_dir, temp_ods]

    subprocess.run(command2, capture_output=True, text=True, check=False)

    

    ret = False

    temp_xlsx = os.path.join(temp_dir, "temp_sanitize_input.xlsx")

    if os.path.exists(temp_xlsx):

        import openpyxl

        try:

            wb = openpyxl.load_workbook(temp_xlsx)

            if len(wb.sheetnames) > 0:

                shutil.copy(temp_xlsx, input_path)

                ret = True

        except Exception:

            pass

            

    # Cleanup temp files

    for f in [temp_input, temp_ods, temp_xlsx]:

        if os.path.exists(f):

            try: os.remove(f)

            except: pass

            

    return ret



def find_signatures_excel(file_path):

    data = load_spreadsheet_data(file_path)

    signatures = []

    for sheet_name, rows in data.items():

        if sheet_name.lower().endswith("_digitalsign") or sheet_name == "Digital Signature":

            if len(rows) > 1 and len(rows[1]) > 0:

                sig = rows[1][0]

                if sig and str(sig).strip() != "":

                    signatures.append(str(sig).strip())

    return signatures



def generate_hash_from_excel(file_path):

    data = load_spreadsheet_data(file_path)

    collected = []

    for sheet_name in sorted(data.keys()):

        if sheet_name.lower().endswith("_digitalsign") or sheet_name == "Digital Signature":

            continue

        rows = data[sheet_name]

        collected.append(f"---SHEET:{sheet_name}---")

        for row in rows:

            if row is None:

                continue

            for value in row:

                if value is None:

                    value = ""

                value = str(value).strip()

                value = value.replace("\r", "")

                value = value.replace("\n", " ")

                try:

                    numeric = float(value)

                    if numeric.is_integer():

                        value = str(int(numeric))

                except:

                    pass

                collected.append(value)

    content = "|".join(collected)

    hash_value = hashlib.sha256(content.encode("utf-8")).hexdigest()

    return hash_value





# =========================================================

# PDF SUPPORT HELPERS

# =========================================================



def find_signatures_pdf(file_path):

    signatures = []

    try:

        with open(file_path, "rb") as f:

            reader = PyPDF2.PdfReader(f)

            for page in reader.pages:

                text = page.extract_text()

                if not text or not text.strip().startswith("Digital Signature"):

                    continue

                

                lines = text.split('\n')

                sig_lines = []

                capture = False

                for line in lines:

                    line = line.strip()

                    if line.startswith("Signature:"):

                        capture = True

                        continue

                    elif line.startswith("Signed by"):

                        capture = False

                        break

                    if capture and line:

                        sig_lines.append(line)

                

                if sig_lines:

                    signature = "".join(sig_lines).strip()

                    if signature:

                        signatures.append(signature)

        return signatures

    except Exception as e:

        print(f"Error extracting signature: {e}")

        return []



def generate_hash_from_pdf(file_path):

    collected = []

    with open(file_path, "rb") as f:

        reader = PyPDF2.PdfReader(f)

        for i, page in enumerate(reader.pages):

            text = page.extract_text()

            if text:

                text = text.replace("\r", "")

                text = text.replace("\n", " ")

                

                if text.strip().startswith("Digital Signature"):

                    continue

                    

                footer_pattern = r"signed by\s+[^\[\]]{1,30}?\[[a-f0-9]{16}\]"

                text = re.sub(footer_pattern, "", text, flags=re.IGNORECASE)

                

                if i == 0:

                    pattern = r"CRC:\s*[a-f0-9]{16}\s*Signed by\s+.*?\s+on\s*\d{2}-\d{2}-\d{4}\s*Approved(?: by .*?)?(?=\s*CRC:|$)"

                    text = re.sub(pattern, "", text, flags=re.IGNORECASE)

                    

                collected.append(f"---PAGE:{i}---")

                collected.append(text.strip())

                

    content = "|".join(collected)

    hash_value = hashlib.sha256(content.encode("utf-8")).hexdigest()

    return hash_value





# =========================================================

# VERIFY SIGNATURE

# =========================================================



def verify_signature(hash_value, signature_b64, public_key_text):

    try:

        signature = base64.b64decode(signature_b64)

        public_key = serialization.load_pem_public_key(public_key_text.encode())

        public_key.verify(

            signature,

            hash_value.encode(),

            padding.PKCS1v15(),

            hashes.SHA256()

        )

        return True

    except InvalidSignature:

        return False

    except Exception as e:

        print(f"Error during verification: {e}")

        return False





# =========================================================

# STREAMLIT UI & LOGIC

# =========================================================



def process_single_verify_file(doc_path, public_key_text):

    file_ext = os.path.splitext(doc_path)[1].lower()

    

    if file_ext == ".pdf":

        signatures = find_signatures_pdf(doc_path)

        if not signatures:

            return False, -1, "Digital signature not found in the PDF.", 0

        hash_value = generate_hash_from_pdf(doc_path)



    elif file_ext in [".xlsx", ".ods"]:

        with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:

            with open(doc_path, "rb") as f:

                tmp.write(f.read())

            temp_path = tmp.name



        if file_ext == ".xlsx":

            import openpyxl

            wb_test = openpyxl.load_workbook(temp_path)

            if len(wb_test.sheetnames) == 0:

                sanitize_xlsx(temp_path)



        try:

            signatures = find_signatures_excel(temp_path)

            if not signatures:

                os.remove(temp_path)

                return False, -1, "Digital signature not found.", 0

            hash_value = generate_hash_from_excel(temp_path)

        finally:

            if os.path.exists(temp_path):

                os.remove(temp_path)

    else:

        return False, -1, "Unsupported file type.", 0



    valid = False

    verified_index = -1

    for idx, sig in enumerate(signatures):

        if verify_signature(hash_value, sig, public_key_text):

            valid = True

            verified_index = idx + 1

            break



    if valid:

        return True, verified_index, "", len(signatures)

    else:

        return False, -1, "SIGNATURE INVALID (File contents were modified or wrong public key was used.)", len(signatures)






st.title("Universal Digital Signature Verifier")

# ─────────────────────────────────────────────
# SECTION 1: Verify Single Document
# ─────────────────────────────────────────────
st.header("1. Verify Single Document")

uploaded_doc = st.file_uploader(
    "Upload Signed Document (PDF, XLSX, ODS)",
    type=["pdf", "xlsx", "ods"],
    key="single_verify_doc"
)
uploaded_key = st.file_uploader(
    "Upload Public Key (.pem)",
    type=["pem"],
    key="single_verify_key"
)

if st.button("Verify Document"):
    if not uploaded_doc:
        st.error("Please upload a signed document.")
    elif not uploaded_key:
        st.error("Please upload the public key.")
    else:
        try:
            public_key_text = uploaded_key.read().decode("utf-8")
            file_ext = os.path.splitext(uploaded_doc.name)[1].lower()

            with st.spinner("Verifying document, please wait..."):
                with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:
                    tmp.write(uploaded_doc.read())
                    temp_path = tmp.name

                try:
                    valid, verified_index, error_msg, num_signatures = process_single_verify_file(temp_path, public_key_text)

                    if valid:
                        if file_ext in [".xlsx", ".ods"] and num_signatures > 1:
                            st.success(f"SIGNATURE VALID\n\nFile is authentic and unchanged.\n\n(Verified matching signature #{verified_index} of {num_signatures} found in document)")
                        else:
                            st.success("SIGNATURE VALID\n\nFile is authentic and unchanged.")
                    else:
                        st.error(f"SIGNATURE INVALID\n\n{error_msg}")
                finally:
                    if os.path.exists(temp_path):
                        os.remove(temp_path)

        except Exception as e:
            st.error(f"Error: {str(e)}")

st.divider()

# ─────────────────────────────────────────────
# SECTION 2: Verify Multiple Documents (Batch)
# ─────────────────────────────────────────────
st.header("2. Verify Multiple Documents (Batch Process)")
st.write("Upload multiple signed documents and your public key to verify them all at once.")

uploaded_batch_docs = st.file_uploader(
    "Upload Signed Documents to Verify (select multiple files)",
    type=["pdf", "xlsx", "ods"],
    accept_multiple_files=True,
    key="batch_verify_docs"
)
uploaded_batch_key = st.file_uploader(
    "Upload Public Key (.pem) for Batch",
    type=["pem"],
    key="batch_verify_key"
)

if st.button("Verify All Uploaded Documents"):
    if not uploaded_batch_docs:
        st.error("Please upload at least one document.")
    elif not uploaded_batch_key:
        st.error("Please upload the public key.")
    else:
        try:
            public_key_text = uploaded_batch_key.read().decode("utf-8")
            supported_exts = [".pdf", ".xlsx", ".ods"]
            success_files = []
            failed_files = []

            with st.spinner(f"Verifying {len(uploaded_batch_docs)} file(s), please wait..."):
                for uploaded_doc_batch in uploaded_batch_docs:
                    filename = uploaded_doc_batch.name
                    file_ext = os.path.splitext(filename)[1].lower()
                    if file_ext not in supported_exts:
                        continue
                    temp_path = None
                    try:
                        with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:
                            tmp.write(uploaded_doc_batch.read())
                            temp_path = tmp.name
                        valid, verified_index, error_msg, num_signatures = process_single_verify_file(temp_path, public_key_text)
                        if valid:
                            success_files.append(filename)
                        else:
                            failed_files.append(f"{filename}: {error_msg}")
                    except Exception as file_e:
                        failed_files.append(f"{filename}: {str(file_e)}")
                    finally:
                        if temp_path and os.path.exists(temp_path):
                            os.remove(temp_path)

            total = len(success_files) + len(failed_files)
            if success_files:
                st.success(f"VALID: {len(success_files)} of {total} file(s) are AUTHENTIC and unchanged:")
                for name in success_files:
                    st.write(f"  [VALID] {name}")
            if failed_files:
                st.error(f"INVALID: {len(failed_files)} of {total} file(s) FAILED verification:")
                for msg in failed_files:
                    st.write(f"  [FAILED] {msg}")
            if not success_files and not failed_files:
                st.info("No supported files (PDF, XLSX, ODS) found in the upload.")

        except Exception as e:
            st.error(f"Batch verification error: {str(e)}")

