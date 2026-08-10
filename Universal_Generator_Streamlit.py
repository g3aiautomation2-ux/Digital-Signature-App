import streamlit as st

import hashlib

import base64

import os

import tempfile

import io

import subprocess

import shutil

import re

from datetime import datetime



# PDF Imports

import PyPDF2

from reportlab.pdfgen import canvas

from reportlab.lib.pagesizes import letter



# Cryptography Imports

from cryptography.hazmat.primitives.asymmetric import rsa, padding

from cryptography.hazmat.primitives import serialization, hashes



# Excel/ODS Imports

from openpyxl import load_workbook, Workbook

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from openpyxl.utils import get_column_letter

from odf.opendocument import load as load_ods

from odf.table import Table, TableRow

from odf.text import P

from odf import teletype





# =========================================================

# ODS SUPPORT HELPERS

# =========================================================



def is_ods_file(file_path):

    return os.path.splitext(str(file_path))[1].lower() == ".ods"



def safe_get_attribute(elem, attr_name):

    if not hasattr(elem, "allowed_attributes"):

        return None

    try:

        allowed = elem.allowed_attributes()

        if allowed:

            allowed_args = [a[1].lower().replace("-", "") for a in allowed]

            if attr_name.lower().replace("-", "") not in allowed_args:

                return None

        return elem.getAttribute(attr_name)

    except Exception:

        return None



def sanitize_color(color_str):

    if not color_str:

        return None

    color_str = str(color_str).lstrip("#").strip().upper()

    if color_str in ("TRANSPARENT", "NONE", ""):

        return None

    if len(color_str) == 3:

        color_str = "".join(c * 2 for c in color_str)

    if len(color_str) in (6, 8) and all(c in "0123456789ABCDEF" for c in color_str):

        return color_str

    return None



def sanitize_sheet_name(name, existing_names):

    if not name:

        name = "Sheet"

    for ch in [":", "\\", "/", "?", "*", "[", "]"]:

        name = name.replace(ch, "_")

    name = name[:31].strip()

    if not name:

        name = "Sheet"



    base_name = name

    counter = 1

    while name in existing_names:

        suffix = f"_{counter}"

        name = base_name[: 31 - len(suffix)] + suffix

        counter += 1



    existing_names.add(name)

    return name



def get_ods_cell_text(cell):

    paragraphs = []

    for p in cell.getElementsByType(P):

        txt = teletype.extractText(p)

        if txt is None:

            txt = ""

        paragraphs.append(txt)

    if paragraphs:

        return "\n".join(paragraphs)

    value = safe_get_attribute(cell, "value")

    if value is not None:

        return str(value)

    return ""



def load_ods_data(file_path):

    doc = load_ods(file_path)

    data = {}

    for table in doc.spreadsheet.getElementsByType(Table):

        sheet_name = safe_get_attribute(table, "name") or "Sheet"

        rows = []

        for row in table.getElementsByType(TableRow):

            values = []

            if hasattr(row, "childNodes"):

                for child in row.childNodes:

                    tag = getattr(child, "tagName", "")

                    if tag in ("table:table-cell", "table:covered-table-cell"):

                        repeat_str = safe_get_attribute(child, "numbercolumnsrepeated")

                        repeat = int(repeat_str) if repeat_str else 1

                        cell_value = get_ods_cell_text(child) if tag == "table:table-cell" else ""

                        for _ in range(repeat):

                            values.append(cell_value)

            while values and values[-1] == "":

                values.pop()

            row_repeat_str = safe_get_attribute(row, "numberrowsrepeated")

            row_repeat = int(row_repeat_str) if row_repeat_str else 1

            for _ in range(row_repeat):

                rows.append(values.copy())

        data[sheet_name] = rows

    return data



def load_spreadsheet_data(file_path):

    if is_ods_file(file_path):

        return load_ods_data(file_path)

    wb = load_workbook(file_path, data_only=False)

    data = {}

    for sheet_name in wb.sheetnames:

        ws = wb[sheet_name]

        rows = []

        for row in ws.iter_rows(min_row=1, min_col=1):

            row_vals = [cell.value for cell in row]

            while row_vals and (row_vals[-1] is None or str(row_vals[-1]).strip() == ""):

                row_vals.pop()

            rows.append(row_vals)

        while rows and len(rows[-1]) == 0:

            rows.pop()

        data[sheet_name] = rows

    return data



def generate_hash_from_excel(file_path):

    data = load_spreadsheet_data(file_path)

    collected = []

    for sheet_name in sorted(data.keys()):

        if sheet_name.lower().endswith("_digitalsign") or sheet_name == "Digital Signature":

            continue

        rows = data[sheet_name]

        collected.append(f"---SHEET:{sheet_name}---")

        for row in rows:

            if row is None:

                continue

            for value in row:

                if value is None:

                    value = ""

                value = str(value).strip()

                value = value.replace("\r", "")

                value = value.replace("\n", " ")

                try:

                    numeric = float(value)

                    if numeric.is_integer():

                        value = str(int(numeric))

                except:

                    pass

                collected.append(value)

    content = "|".join(collected)

    hash_value = hashlib.sha256(content.encode("utf-8")).hexdigest()

    return hash_value



def _convert_with_excel(input_path, output_path):

    import win32com.client

    excel = None

    workbook = None

    try:

        excel = win32com.client.Dispatch("Excel.Application")

        excel.Visible = False

        excel.DisplayAlerts = False

        workbook = excel.Workbooks.Open(os.path.abspath(input_path), ReadOnly=True)

        workbook.SaveAs(os.path.abspath(output_path), FileFormat=51)

    finally:

        if workbook is not None:

            try:

                workbook.Close(SaveChanges=False)

            except Exception:

                pass

        if excel is not None:

            try:

                excel.Quit()

            except Exception:

                pass



def _find_libreoffice_command():

    candidates = [

        shutil.which("soffice"),

        shutil.which("libreoffice"),

        r"C:\Program Files\LibreOffice\program\soffice.exe",

        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",

        r"C:\Program Files\OpenOffice\program\soffice.exe",

        r"C:\Program Files (x86)\OpenOffice\program\soffice.exe"

    ]

    for candidate in candidates:

        if candidate and os.path.exists(candidate):

            return candidate

    return None



def _convert_with_libreoffice(input_path, output_path):

    libreoffice_cmd = _find_libreoffice_command()

    if not libreoffice_cmd:

        raise RuntimeError("LibreOffice is not available on this system.")

    command = [

        libreoffice_cmd, "--headless", "--convert-to", "xlsx",

        "--outdir", os.path.dirname(os.path.abspath(output_path)),

        os.path.abspath(input_path)

    ]

    result = subprocess.run(command, capture_output=True, text=True, check=False)

    if result.returncode != 0:

        raise RuntimeError(f"LibreOffice conversion failed: {result.stderr.strip() or result.stdout.strip()}")

    converted_name = os.path.splitext(os.path.basename(input_path))[0] + ".xlsx"

    converted_path = os.path.join(os.path.dirname(os.path.abspath(output_path)), converted_name)

    if not os.path.exists(converted_path):

        raise RuntimeError("LibreOffice conversion did not produce an XLSX file.")

    os.replace(converted_path, output_path)



def sanitize_xlsx(input_path):

    libreoffice_cmd = _find_libreoffice_command()

    if not libreoffice_cmd:

        return False

        

    temp_dir = os.path.dirname(input_path)

    temp_name = "temp_sanitize_input.xlsx"

    temp_input = os.path.join(temp_dir, temp_name)

    shutil.copy(input_path, temp_input)

    

    command1 = [

        libreoffice_cmd, "--headless", "--convert-to", "ods",

        "--outdir", temp_dir, temp_input

    ]

    subprocess.run(command1, capture_output=True, text=True, check=False)

    

    temp_ods = os.path.join(temp_dir, "temp_sanitize_input.ods")

    if not os.path.exists(temp_ods):

        return False

        

    command2 = [

        libreoffice_cmd, "--headless", "--convert-to", "xlsx",

        "--outdir", temp_dir, temp_ods

    ]

    subprocess.run(command2, capture_output=True, text=True, check=False)

    

    ret = False

    temp_xlsx = os.path.join(temp_dir, "temp_sanitize_input.xlsx")

    if os.path.exists(temp_xlsx):

        import openpyxl

        try:

            wb = openpyxl.load_workbook(temp_xlsx)

            if len(wb.sheetnames) > 0:

                shutil.copy(temp_xlsx, input_path)

                ret = True

        except Exception:

            pass

            

    # Cleanup temp files

    for f in [temp_input, temp_ods, temp_xlsx]:

        if os.path.exists(f):

            try: os.remove(f)

            except: pass

            

    return ret



def parse_odf_styles(doc):

    styles = {}

    sources = []

    if hasattr(doc, "automaticstyles") and doc.automaticstyles:

        sources.extend(list(doc.automaticstyles.childNodes))

    if hasattr(doc, "styles") and doc.styles:

        sources.extend(list(doc.styles.childNodes))

    for s in sources:

        name = safe_get_attribute(s, "name")

        if not name:

            continue

        style_info = {"bg_color": None, "bold": False, "align": None, "color": None, "border": False}

        if hasattr(s, "childNodes"):

            for child in s.childNodes:

                if not hasattr(child, "attributes"):

                    continue

                for (ns, attr), val in child.attributes.items():

                    if attr == "background-color" and val != "transparent":

                        style_info["bg_color"] = val.lstrip("#").upper()

                    elif attr == "font-weight" and val in ("bold", "700", "800", "900"):

                        style_info["bold"] = True

                    elif attr == "text-align":

                        style_info["align"] = val

                    elif attr == "color":

                        style_info["color"] = val.lstrip("#").upper()

                    elif "border" in attr and val != "none":

                        style_info["border"] = True

        styles[name] = style_info

    return styles



def _copy_ods_merges_to_ws(table, ws):

    from openpyxl.utils import get_column_letter as gcl

    row_idx = 1

    for row in table.getElementsByType(TableRow):

        rows_repeated = int(safe_get_attribute(row, "numberrowsrepeated") or 1)

        col_idx = 1

        if hasattr(row, "childNodes"):

            for child in row.childNodes:

                tag = getattr(child, "tagName", "")

                if tag not in ("table:table-cell", "table:covered-table-cell"):

                    continue

                cols_repeated = int(safe_get_attribute(child, "numbercolumnsrepeated") or 1)

                if tag == "table:table-cell":

                    cols_spanned = int(safe_get_attribute(child, "numbercolumnsspanned") or 1)

                    rows_spanned = int(safe_get_attribute(child, "numberrowsspanned") or 1)

                    if cols_spanned > 1 or rows_spanned > 1:

                        r1, c1 = row_idx, col_idx

                        r2 = row_idx + rows_spanned - 1

                        c2 = col_idx + cols_spanned - 1

                        region = f"{gcl(c1)}{r1}:{gcl(c2)}{r2}"

                        try:

                            ws.merge_cells(region)

                        except Exception:

                            pass

                    col_idx += cols_repeated * (int(safe_get_attribute(child, "numbercolumnsspanned") or 1))

                else:

                    col_idx += cols_repeated

        row_idx += rows_repeated



def _convert_with_python(input_path, output_path):

    doc = load_ods(input_path)

    styles = parse_odf_styles(doc)

    wb = Workbook()

    existing_sheet_names = set()

    first_sheet = True

    thin_border = Border(

        left=Side(style="thin", color="D0D0D0"),

        right=Side(style="thin", color="D0D0D0"),

        top=Side(style="thin", color="D0D0D0"),

        bottom=Side(style="thin", color="D0D0D0")

    )

    for table in doc.spreadsheet.getElementsByType(Table):

        raw_name = safe_get_attribute(table, "name")

        sheet_name = sanitize_sheet_name(raw_name, existing_sheet_names)

        if first_sheet:

            ws = wb.active

            ws.title = sheet_name

            first_sheet = False

        else:

            ws = wb.create_sheet(sheet_name)

        row_idx = 1

        for row in table.getElementsByType(TableRow):

            row_repeat_str = safe_get_attribute(row, "numberrowsrepeated")

            row_repeat = int(row_repeat_str) if row_repeat_str else 1

            cells_data = []

            has_row_data = False

            if hasattr(row, "childNodes"):

                for child in row.childNodes:

                    tag = getattr(child, "tagName", "")

                    if tag in ("table:table-cell", "table:covered-table-cell"):

                        repeat_str = safe_get_attribute(child, "numbercolumnsrepeated")

                        repeat = int(repeat_str) if repeat_str else 1

                        if tag == "table:table-cell":

                            cell_value = get_ods_cell_text(child)

                            style_name = safe_get_attribute(child, "stylename")

                            style_info = styles.get(style_name, {}) if style_name else {}

                        else:

                            cell_value = ""

                            style_info = {}

                        for _ in range(repeat):

                            cells_data.append((cell_value, style_info))

                            if cell_value != "":

                                has_row_data = True

            while cells_data and cells_data[-1][0] == "":

                cells_data.pop()

            if not has_row_data and row_idx > 1 and row_repeat > 5:

                break

            for _ in range(row_repeat):

                for col_idx, (val, st_info) in enumerate(cells_data, 1):

                    c = ws.cell(row=row_idx, column=col_idx)

                    if isinstance(val, str):

                        val = val.replace("\r\n", "\n").replace("\r", "\n")

                    c.value = val

                    try:

                        bg = sanitize_color(st_info.get("bg_color"))

                        if bg:

                            c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

                        is_bold = st_info.get("bold", False)

                        font_color = sanitize_color(st_info.get("color"))

                        if is_bold or font_color:

                            c.font = Font(bold=is_bold, color=font_color if font_color else "000000")

                        align = st_info.get("align")

                        if align in ("left", "center", "right", "justify"):

                            c.alignment = Alignment(horizontal=align)

                        if st_info.get("border"):

                            c.border = thin_border

                    except Exception:

                        pass

                row_idx += 1

        try:

            _copy_ods_merges_to_ws(table, ws)

        except Exception:

            pass

        try:

            format_worksheet_layout(ws)

        except Exception:

            pass

    if len(wb.sheetnames) == 0:

        wb.create_sheet("Sheet1")

    wb.save(output_path)



def convert_ods_to_xlsx(input_path, output_path):

    try:

        _convert_with_excel(input_path, output_path)

        return

    except Exception:

        pass

    try:

        _convert_with_libreoffice(input_path, output_path)

        return

    except Exception:

        pass

    try:

        _convert_with_python(input_path, output_path)

        return

    except Exception as exc:

        raise RuntimeError(f"Unable to convert ODS file to XLSX: {exc}") from exc



def format_worksheet_layout(ws):

    thin_border = Border(

        left=Side(style="thin", color="D0D0D0"),

        right=Side(style="thin", color="D0D0D0"),

        top=Side(style="thin", color="D0D0D0"),

        bottom=Side(style="thin", color="D0D0D0")

    )

    merged_slave_coords = set()

    for merged_range in ws.merged_cells.ranges:

        for row_idx in range(merged_range.min_row, merged_range.max_row + 1):

            for col_idx in range(merged_range.min_col, merged_range.max_col + 1):

                if row_idx == merged_range.min_row and col_idx == merged_range.min_col:

                    continue

                merged_slave_coords.add((row_idx, col_idx))

    for col in ws.columns:

        max_len = 0

        col_letter = get_column_letter(col[0].column)

        for cell in col:

            if (cell.row, cell.column) in merged_slave_coords:

                continue

            val = str(cell.value or "")

            if "\n" in val:

                lines = val.split("\n")

                line_len = max(len(l) for l in lines) if lines else 0

            else:

                line_len = len(val)

            if line_len > max_len:

                max_len = line_len

        if max_len > 0:

            calculated_w = max(max_len + 6, 18)

            ws.column_dimensions[col_letter].width = min(calculated_w, 55)

    for row in ws.iter_rows():

        if not any(cell.value is not None and str(cell.value).strip() != "" for cell in row):

            continue

        row_num = row[0].row

        max_lines = 1

        for cell in row:

            if (cell.row, cell.column) in merged_slave_coords:

                continue

            if cell.value is not None:

                val = str(cell.value)

                col_letter = get_column_letter(cell.column)

                col_width = ws.column_dimensions[col_letter].width or 18

                char_per_line = max(int(col_width * 1.2), 10)

                wrapped_lines = max(len(val) // char_per_line + 1, val.count("\n") + 1)

                if wrapped_lines > max_lines:

                    max_lines = wrapped_lines

        row_height = max(24, min(max_lines * 18, 120))

        ws.row_dimensions[row_num].height = row_height

        for cell in row:

            if (cell.row, cell.column) in merged_slave_coords:

                continue

            if cell.value is not None and str(cell.value).strip() != "":

                existing_h = cell.alignment.horizontal if cell.alignment else None

                cell.alignment = Alignment(

                    horizontal=existing_h if existing_h and existing_h != "general" else "left",

                    vertical="center",

                    wrap_text=True

                )

                if not cell.border or getattr(cell.border, "left", None) is None or getattr(cell.border.left, "style", None) is None:

                    cell.border = thin_border



def store_signature_excel(file_path, signature_b64, approver_name):

    wb = load_workbook(file_path)

    sheet_name = f"{approver_name}_digitalsign"

    

    if sheet_name in wb.sheetnames:

        del wb[sheet_name]

        

    ws = wb.create_sheet(sheet_name)

    start_row = 1

            

    crc_value = hashlib.sha256(signature_b64.encode()).hexdigest()[:16]

    ws.cell(row=start_row, column=1, value=crc_value)

    ws.cell(row=start_row+1, column=1, value=signature_b64)

    current_date = datetime.now().strftime("%d-%m-%Y")

    ws.cell(row=start_row+2, column=1, value=f"Approved by {approver_name} on {current_date}")

    try:

        format_worksheet_layout(ws)

    except Exception:

        pass

    wb.save(file_path)



# =========================================================

# PDF SUPPORT HELPERS

# =========================================================



def generate_hash_from_pdf(file_path):

    collected = []

    with open(file_path, "rb") as f:

        reader = PyPDF2.PdfReader(f)

        for i, page in enumerate(reader.pages):

            text = page.extract_text()

            if text:

                text = text.replace("\r", "")

                text = text.replace("\n", " ")

                

                if text.strip().startswith("Digital Signature"):

                    continue

                    

                footer_pattern = r"signed by\s+[^\[\]]{1,30}?\[[a-f0-9]{16}\]"

                text = re.sub(footer_pattern, "", text, flags=re.IGNORECASE)

                

                if i == 0:

                    pattern = r"CRC:\s*[a-f0-9]{16}\s*Signed by\s+.*?\s+on\s*\d{2}-\d{2}-\d{4}\s*Approved(?: by .*?)?(?=\s*CRC:|$)"

                    text = re.sub(pattern, "", text, flags=re.IGNORECASE)

                    

                collected.append(f"---PAGE:{i}---")

                collected.append(text.strip())

                

    content = "|".join(collected)

    hash_value = hashlib.sha256(content.encode("utf-8")).hexdigest()

    return hash_value



def store_signature_pdf(file_path, signature_b64, approver_text, approver_name):

    crc_value = hashlib.sha256(signature_b64.encode()).hexdigest()[:16]

    current_date = datetime.now().strftime("%d-%m-%Y")

    

    # Try to load Calibri, fallback to Helvetica

    from reportlab.pdfbase import pdfmetrics

    from reportlab.pdfbase.ttfonts import TTFont

    import os

    font_name = "Helvetica"

    try:

        calibri_path = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts', 'calibri.ttf')

        if os.path.exists(calibri_path):

            pdfmetrics.registerFont(TTFont('Calibri', calibri_path))

            font_name = "Calibri"

    except:

        pass

        

    existing_signatures = 0

    with open(file_path, "rb") as f:

        reader = PyPDF2.PdfReader(f)

        for page in reader.pages:

            text = page.extract_text()

            if text and text.strip().startswith("Digital Signature"):

                existing_signatures += 1



    packet = io.BytesIO()

    can = canvas.Canvas(packet, pagesize=letter)

    can.setFont(font_name, 14)

    can.drawString(50, 750, "Digital Signature")

    can.setFont(font_name, 10)

    

    can.drawString(50, 710, f"CRC: {crc_value}")

    can.drawString(50, 690, "Signature:")

    

    sig_lines = [signature_b64[i:i+64] for i in range(0, len(signature_b64), 64)]

    y = 675

    for line in sig_lines:

        can.drawString(70, y, line)

        y -= 15

        

    y -= 15

    can.drawString(50, y, f"Signed by {approver_name} on {current_date}")

    can.save()

    packet.seek(0)

    new_pdf = PyPDF2.PdfReader(packet)

    

    stamp_y = 750 - (existing_signatures * 55)

    stamp_packet = io.BytesIO()

    stamp_can = canvas.Canvas(stamp_packet, pagesize=letter)

    stamp_can.setStrokeColorRGB(0, 0.6, 0)

    stamp_can.setFillColorRGB(0, 0.6, 0)

    stamp_can.setLineWidth(1.2)

    stamp_can.roundRect(430, stamp_y, 130, 45, 5)

    stamp_can.setFont(font_name, 8)

    stamp_can.drawString(435, stamp_y + 30, f"CRC: {crc_value}")

    stamp_can.drawString(435, stamp_y + 18, f"Signed by {approver_name} on {current_date}")

    stamp_can.drawString(435, stamp_y + 6, approver_text)

    stamp_can.save()

    stamp_packet.seek(0)

    stamp_pdf = PyPDF2.PdfReader(stamp_packet)

    stamp_page = stamp_pdf.pages[0]

    

    footer_y = 50 - (existing_signatures * 15)

    footer_packet = io.BytesIO()

    footer_can = canvas.Canvas(footer_packet, pagesize=letter)

    footer_can.setFont(font_name, 10)

    footer_can.setFillColorRGB(0.5, 0.5, 0.5)

    footer_text = f"Signed by {approver_name} [{crc_value}]"

    footer_can.drawCentredString(306, footer_y, footer_text)

    footer_can.save()

    footer_packet.seek(0)

    footer_pdf = PyPDF2.PdfReader(footer_packet)

    footer_page = footer_pdf.pages[0]

    

    with open(file_path, "rb") as f:

        existing_pdf = PyPDF2.PdfReader(f)

        output = PyPDF2.PdfWriter()

        for i, page in enumerate(existing_pdf.pages):

            page.merge_page(footer_page)

            if i == 0:

                page.merge_page(stamp_page)

            output.add_page(page)

        output.add_page(new_pdf.pages[0])

        with open(file_path + ".tmp", "wb") as outputStream:

            output.write(outputStream)

    shutil.move(file_path + ".tmp", file_path)



# =========================================================

# COMMON SIGNATURE HELPERS

# =========================================================



def generate_new_keys():

    private_key = rsa.generate_private_key(public_exponent=65537, key_size=2048)

    public_key = private_key.public_key()

    private_key_pem = private_key.private_bytes(

        encoding=serialization.Encoding.PEM,

        format=serialization.PrivateFormat.PKCS8,

        encryption_algorithm=serialization.NoEncryption()

    )

    public_key_pem = public_key.public_bytes(

        encoding=serialization.Encoding.PEM,

        format=serialization.PublicFormat.SubjectPublicKeyInfo

    )

    return private_key_pem, public_key_pem



def sign_hash(hash_value, private_key):

    signature = private_key.sign(

        hash_value.encode(),

        padding.PKCS1v15(),

        hashes.SHA256()

    )

    return base64.b64encode(signature).decode()



# =========================================================

# STREAMLIT UI

# =========================================================






st.title("Universal Digital Signature Generator")



st.header("1. Generate New Keys")

if st.button("Generate Key Pair"):

    try:

        import zipfile

        priv_pem, pub_pem = generate_new_keys()

        

        # Create an in-memory zip file

        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zf:

            zf.writestr("private_key.pem", priv_pem)

            zf.writestr("public_key.pem", pub_pem)

            

        st.success("Keys generated! Download them as a ZIP folder below.")

        st.download_button("Download Keys (ZIP)", zip_buffer.getvalue(), "keys.zip", mime="application/zip")

    except Exception as e:

        st.error(f"Error generating keys: {e}")



st.divider()



st.header("2. Sign Document")

uploaded_doc = st.file_uploader("Upload Document (PDF, XLSX, ODS)", type=["pdf", "xlsx", "ods"])

uploaded_key = st.file_uploader("Upload Private Key (.pem)", type=["pem"])



if st.button("Sign Document"):

    if not uploaded_doc:

        st.error("Please upload a document.")

    elif not uploaded_key:

        st.error("Please upload your private key.")

    else:

        try:

            # Check key filename format

            basename = uploaded_key.name

            name_part_full = os.path.splitext(basename)[0]

            match = re.search(r'_(?i:private(_key)?)$', name_part_full)

            if not match or match.start() == 0:

                st.error("The private key filename must be in the format '[Name]_Private.pem'.\\n\\nExample: Raksha_Private.pem")

            else:

                name_part = name_part_full[:match.start()]

                approver_text = f"Approved by {name_part}"

                

                # Load private key

                private_key = serialization.load_pem_private_key(uploaded_key.read(), password=None)

                

                file_ext = os.path.splitext(uploaded_doc.name)[1].lower()

                

                with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:

                    tmp.write(uploaded_doc.read())

                    temp_path = tmp.name

                    

                if file_ext == ".pdf":

                    hash_value = generate_hash_from_pdf(temp_path)

                    signature_b64 = sign_hash(hash_value, private_key)

                    store_signature_pdf(temp_path, signature_b64, approver_text, name_part)

                    output_name = uploaded_doc.name.replace(".pdf", "_Signed.pdf")

                    

                elif file_ext in [".xlsx", ".ods"]:

                    working_path = temp_path

                    converted_tmp = False

                    

                    if file_ext == ".xlsx":

                        # Some XLSX files (especially from LibreOffice) have strict XML that openpyxl fails to read,

                        # causing openpyxl to see 0 sheets and delete the original data on save. 

                        # We sanitize by converting to ODS and back.

                        wb_test = load_workbook(temp_path)

                        if len(wb_test.sheetnames) == 0:

                            sanitize_xlsx(temp_path)



                    if file_ext == ".ods":

                        working_path = temp_path + ".converted.xlsx"

                        convert_ods_to_xlsx(temp_path, working_path)

                        converted_tmp = True

                        

                    hash_value = generate_hash_from_excel(working_path)

                    signature_b64 = sign_hash(hash_value, private_key)

                    

                    output_name = os.path.splitext(uploaded_doc.name)[0] + "_Signed.xlsx"

                    store_signature_excel(working_path, signature_b64, name_part)

                    

                    if converted_tmp:

                        # move the converted file back to temp_path for reading

                        shutil.copy(working_path, temp_path)

                        os.remove(working_path)

                

                with open(temp_path, "rb") as f:

                    final_data = f.read()

                    

                st.success("Document successfully signed! Download below:")

                st.download_button(label=f"Download {output_name}", data=final_data, file_name=output_name)

                

                os.remove(temp_path)

                

        except Exception as e:

            st.error(f"Error: {str(e)}")



st.divider()



st.header("3. Sign Folder (Batch Process)")

st.write("Enter the absolute path to a folder on your computer to process all supported documents inside it.")



folder_path = st.text_input("Folder Path (e.g., C:\\Users\\Name\\Downloads\\Documents)")

uploaded_batch_key = st.file_uploader("Upload Private Key (.pem) for Batch", type=["pem"], key="batch_key_uploader")



if st.button("Sign All Documents in Folder"):

    if not folder_path or not os.path.exists(folder_path) or not os.path.isdir(folder_path):

        st.error("Please enter a valid directory path on your computer.")

    elif not uploaded_batch_key:

        st.error("Please upload your private key.")

    else:

        try:

            basename = uploaded_batch_key.name

            name_part_full = os.path.splitext(basename)[0]

            match = re.search(r'_(?i:private(_key)?)$', name_part_full)

            if not match or match.start() == 0:

                st.error("The private key filename must be in the format '[Name]_Private.pem'.\\n\\nExample: Raksha_Private.pem")

            else:

                name_part = name_part_full[:match.start()]

                approver_text = f"Approved by {name_part}"

                

                private_key = serialization.load_pem_private_key(uploaded_batch_key.read(), password=None)

                

                success_count = 0

                error_messages = []

                

                supported_exts = [".pdf", ".xlsx", ".ods"]

                

                output_dir = os.path.join(folder_path, f"Signed_by_{name_part}")

                if not os.path.exists(output_dir):

                    os.makedirs(output_dir)

                    

                with st.spinner("Processing files..."):

                    for filename in os.listdir(folder_path):

                        file_ext = os.path.splitext(filename)[1].lower()

                        if file_ext in supported_exts:

                            doc_path = os.path.join(folder_path, filename)

                            

                            output_filename = os.path.basename(doc_path)

                            if file_ext == ".ods":

                                output_filename = os.path.splitext(output_filename)[0] + "_Signed.xlsx"

                            elif not output_filename.endswith(f"_Signed{file_ext}"):

                                output_filename = os.path.splitext(output_filename)[0] + f"_Signed{file_ext}"

                                

                            output_path = os.path.join(output_dir, output_filename)

                            

                            try:

                                if file_ext == ".pdf":

                                    shutil.copy(doc_path, output_path)

                                    hash_value = generate_hash_from_pdf(output_path)

                                    signature_b64 = sign_hash(hash_value, private_key)

                                    store_signature_pdf(output_path, signature_b64, approver_text, name_part)

                                    

                                elif file_ext in [".xlsx", ".ods"]:

                                    working_path = doc_path

                                    converted_tmp = False

                                    

                                    if file_ext == ".xlsx":

                                        wb_test = load_workbook(working_path)

                                        if len(wb_test.sheetnames) == 0:

                                            sanitize_xlsx(working_path)

            

                                    if file_ext == ".ods":

                                        working_path = doc_path + ".converted.xlsx"

                                        convert_ods_to_xlsx(doc_path, working_path)

                                        converted_tmp = True

                                        

                                    hash_value = generate_hash_from_excel(working_path)

                                    signature_b64 = sign_hash(hash_value, private_key)

                                    

                                    if working_path != output_path:

                                        shutil.copy(working_path, output_path)

                                    

                                    store_signature_excel(output_path, signature_b64, name_part)

                                    

                                    if converted_tmp and os.path.exists(working_path):

                                        os.remove(working_path)

                                        

                                success_count += 1

                                

                            except Exception as file_e:

                                error_messages.append(f"{filename}: {str(file_e)}")

                

                if success_count > 0:

                    st.success(f"Successfully signed {success_count} documents!\n\nThey have been saved directly to your computer at:\n`{output_dir}`")

                elif not error_messages:

                    st.info("No supported files (PDF, XLSX, ODS) found in the selected folder.")

                

                if error_messages:

                    st.warning("Some files encountered errors during processing:")

                    for err in error_messages:

                        st.write(f"- {err}")

                        

        except Exception as e:

            st.error(f"Batch processing error: {str(e)}")

