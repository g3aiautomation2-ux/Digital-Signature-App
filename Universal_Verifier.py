import tkinter as tk
from tkinter import filedialog, messagebox
import hashlib
import base64
import os
import re
import tempfile
import shutil
# PDF Imports
import PyPDF2

# Cryptography Imports
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.asymmetric import padding
from cryptography.hazmat.primitives import hashes
from cryptography.exceptions import InvalidSignature

# Excel/ODS Imports
from openpyxl import load_workbook
from odf.opendocument import load as load_ods
from odf.table import Table, TableRow, TableCell
from odf.text import P
from odf import teletype


# =========================================================
# ODS SUPPORT HELPERS
# =========================================================

def is_ods_file(file_path):
    return os.path.splitext(str(file_path))[1].lower() == ".ods"

def get_ods_cell_text(cell):
    text = teletype.extractText(cell)
    if text is not None:
        text = text.strip()
    if text:
        return text
    value = cell.getAttribute("value")
    if value is not None:
        return str(value)
    return ""

def load_ods_data(file_path):
    doc = load_ods(file_path)
    data = {}
    for table in doc.spreadsheet.getElementsByType(Table):
        sheet_name = table.getAttribute("name")
        rows = []
        for row in table.getElementsByType(TableRow):
            values = []
            for cell in row.getElementsByType(TableCell):
                repeat = cell.getAttribute("numbercolumnsrepeated")
                repeat = int(repeat) if repeat else 1
                cell_value = get_ods_cell_text(cell)
                for _ in range(repeat):
                    values.append(cell_value)
            while values and values[-1] == "":
                values.pop()
            row_repeat = row.getAttribute("numberrowsrepeated")
            row_repeat = int(row_repeat) if row_repeat else 1
            for _ in range(row_repeat):
                rows.append(values.copy())
        data[sheet_name] = rows
    return data

def load_spreadsheet_data(file_path):
    if is_ods_file(file_path):
        return load_ods_data(file_path)
    wb = load_workbook(file_path, data_only=False)
    data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=1, min_col=1):
            row_vals = [cell.value for cell in row]
            while row_vals and (row_vals[-1] is None or str(row_vals[-1]).strip() == ""):
                row_vals.pop()
            rows.append(row_vals)
        while rows and len(rows[-1]) == 0:
            rows.pop()
        data[sheet_name] = rows
    return data

def find_signature_excel(file_path):
    data = load_spreadsheet_data(file_path)
    if "Digital Signature" not in data:
        return None
    rows = data["Digital Signature"]
    if len(rows) < 2 or len(rows[1]) == 0:
        return None
    signature = rows[1][0]
    if signature is None or str(signature).strip() == "":
        return None
    return signature

def generate_hash_from_excel(file_path):
    data = load_spreadsheet_data(file_path)
    collected = []
    for sheet_name in sorted(data.keys()):
        if sheet_name == "Digital Signature":
            continue
        rows = data[sheet_name]
        collected.append(f"---SHEET:{sheet_name}---")
        for row in rows:
            if row is None:
                continue
            for value in row:
                if value is None:
                    value = ""
                value = str(value).strip()
                value = value.replace("\r", "")
                value = value.replace("\n", " ")
                try:
                    numeric = float(value)
                    if numeric.is_integer():
                        value = str(int(numeric))
                except:
                    pass
                collected.append(value)
    content = "|".join(collected)
    hash_value = hashlib.sha256(content.encode("utf-8")).hexdigest()
    return hash_value


# =========================================================
# PDF SUPPORT HELPERS
# =========================================================

def find_signature_pdf(file_path):
    try:
        with open(file_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            num_pages = len(reader.pages)
            if num_pages == 0:
                return None
            last_page = reader.pages[-1]
            text = last_page.extract_text()
            if not text:
                return None
            lines = text.split('\n')
            sig_lines = []
            capture = False
            for line in lines:
                line = line.strip()
                if line.startswith("Signature:"):
                    capture = True
                    continue
                elif line.startswith("Signed by"):
                    capture = False
                    break
                if capture and line:
                    sig_lines.append(line)
            if not sig_lines:
                return None
            signature = "".join(sig_lines).strip()
            if signature == "":
                return None
            return signature
    except Exception as e:
        print(f"Error extracting signature: {e}")
        return None

def generate_hash_from_pdf(file_path):
    collected = []
    with open(file_path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        num_pages = len(reader.pages)
        for i in range(num_pages - 1):
            page = reader.pages[i]
            collected.append(f"---PAGE:{i}---")
            text = page.extract_text()
            if text:
                text = text.replace("\r", "")
                text = text.replace("\n", " ")
                footer_pattern = r"signed by\s+[a-zA-Z0-9_\- ]+\s+\[[a-f0-9]{16}\]"
                text = re.sub(footer_pattern, "", text, flags=re.IGNORECASE)
                if i == 0:
                    pattern = r"CRC:\s*[a-f0-9]{16}\s*Signed by\s+[a-zA-Z0-9_\- ]+\s+on\s*\d{2}-\d{2}-\d{4}\s*Approved(?: by [a-zA-Z0-9_\- ]+)?"
                    text = re.sub(pattern, "", text, flags=re.IGNORECASE)
                collected.append(text.strip())
    content = "|".join(collected)
    hash_value = hashlib.sha256(content.encode("utf-8")).hexdigest()
    return hash_value


# =========================================================
# VERIFY SIGNATURE
# =========================================================

def verify_signature(hash_value, signature_b64, public_key_text):
    try:
        signature = base64.b64decode(signature_b64)
        public_key = serialization.load_pem_public_key(public_key_text.encode())
        public_key.verify(
            signature,
            hash_value.encode(),
            padding.PKCS1v15(),
            hashes.SHA256()
        )
        return True
    except InvalidSignature:
        return False
    except Exception as e:
        print(f"Error during verification: {e}")
        return False


def verify_document_gui():
    try:
        doc_path = filedialog.askopenfilename(
            title="Select Signed Document File",
            filetypes=[("Supported Files", "*.pdf;*.xlsx;*.ods"), ("All Files", "*.*")]
        )
        if not doc_path:
            return

        public_key_path = filedialog.askopenfilename(
            title="Select Public Key",
            filetypes=[("PEM Files", "*.pem")]
        )
        if not public_key_path:
            return

        with open(public_key_path, "r", encoding="utf-8") as f:
            public_key_text = f.read()

        file_ext = os.path.splitext(doc_path)[1].lower()

        if file_ext == ".pdf":
            signature_b64 = find_signature_pdf(doc_path)
            if signature_b64 is None:
                messagebox.showerror("Verification Failed", "Digital signature not found on the last page of the PDF.")
                return
            hash_value = generate_hash_from_pdf(doc_path)

        elif file_ext in [".xlsx", ".ods"]:
            # Need to copy to temp in case it's locked, similar to how app_verifier handled it.
            with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:
                with open(doc_path, "rb") as f:
                    tmp.write(f.read())
                temp_path = tmp.name

            try:
                signature_b64 = find_signature_excel(temp_path)
                if signature_b64 is None:
                    messagebox.showerror("Verification Failed", "Digital signature not found.")
                    os.remove(temp_path)
                    return
                hash_value = generate_hash_from_excel(temp_path)
            finally:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
        else:
            messagebox.showerror("Error", "Unsupported file type.")
            return

        valid = verify_signature(hash_value, signature_b64, public_key_text)

        if valid:
            messagebox.showinfo("Verification Successful", "SIGNATURE VALID\n\nFile is authentic and unchanged.")
        else:
            messagebox.showerror("Verification Failed", "SIGNATURE INVALID\n\nFile contents were modified\nor wrong public key was used.")

    except Exception as e:
        messagebox.showerror("Error", str(e))


# =========================================================
# GUI
# =========================================================
if __name__ == "__main__":
    root = tk.Tk()
    root.title("Universal Digital Signature Verifier")
    root.geometry("500x250")
    tk.Label(root, text="Universal Digital Signature Verifier", font=("Arial", 16, "bold")).pack(pady=25)
    tk.Button(root, text="Verify Signed Document (PDF/XLSX/ODS)", width=35, height=2, command=verify_document_gui).pack(pady=20)
    root.mainloop()
